#!/usr/bin/env python3
from __future__ import annotations

import json
import math
import re
import time
import zipfile
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any
from urllib.parse import quote

import numpy as np
import pandas as pd
import requests
from openpyxl import load_workbook

BASE_DIR = Path('/Users/melodysong/code/phd/battery_ml')
DATA_DIR = BASE_DIR / 'data/zn_ion/batterylife/ZNion'
PARQUET_PATH = BASE_DIR / 'processed/zn_ion_raw.parquet'
SUMMARY_PATH = BASE_DIR / 'processed/znion_batches_summary.json'
BATCHES_TO_DOWNLOAD = ('Batch-1', 'Batch-3')
HF_API_BASE = 'https://huggingface.co/api/datasets/Battery-Life/BatteryLife_Raw/tree/main/'
HF_FILE_BASE = 'https://huggingface.co/datasets/Battery-Life/BatteryLife_Raw/resolve/main/'
TEMP_STATS_URL = (
    'https://raw.githubusercontent.com/Ruifeng-Tan/BatteryLife/main/assets/'
    'Further_details_of_data_statistics.md'
)
TEMP_PREPROCESS_URL = (
    'https://raw.githubusercontent.com/Ruifeng-Tan/BatteryLife/main/process_scripts/'
    'preprocess_ZNion.py'
)
TEMP_AGING_URL = 'https://raw.githubusercontent.com/Ruifeng-Tan/BatteryLife/main/aging_conditions.py'


def _session() -> requests.Session:
    sess = requests.Session()
    sess.headers.update({'User-Agent': 'Mozilla/5.0'})
    return sess


def list_hf_files(batch: str) -> list[dict[str, Any]]:
    url = f'{HF_API_BASE}ZNion/{batch}'
    resp = _session().get(url, timeout=60)
    resp.raise_for_status()
    files = [item for item in resp.json() if item['path'].endswith('.xlsx')]
    files.sort(key=lambda item: item['path'])
    return files


def download_one(item: dict[str, Any]) -> Path:
    rel_path = item['path']
    local_path = DATA_DIR / Path(rel_path).relative_to('ZNion')
    local_path.parent.mkdir(parents=True, exist_ok=True)
    expected_size = int(item.get('size', 0))
    if local_path.exists() and local_path.stat().st_size == expected_size:
        return local_path

    url = HF_FILE_BASE + quote(rel_path)
    sess = _session()
    for attempt in range(3):
        try:
            with sess.get(url, stream=True, timeout=180) as resp:
                resp.raise_for_status()
                with open(local_path, 'wb') as fh:
                    for chunk in resp.iter_content(chunk_size=1024 * 1024):
                        if chunk:
                            fh.write(chunk)
            return local_path
        except Exception:
            if attempt == 2:
                raise
            time.sleep(2 + attempt)
    return local_path


def cell_id_from_name(name: str) -> str:
    stem = Path(name).stem
    dash_match = re.search(r'(\d+-\d+)', stem)
    if dash_match:
        return dash_match.group(1)
    if '我的设备_' in stem:
        left, right = stem.split('我的设备_', 1)
        return f"{left.rstrip('_')}_{right}".replace('__', '_')
    return stem


def normalize_header(row: Any) -> dict[str, int]:
    header: dict[str, int] = {}
    if row is None:
        return header
    for idx, value in enumerate(row):
        if value is None:
            continue
        header[str(value).strip()] = idx
    return header


def read_sheet_rows(ws) -> list[dict[str, Any]]:
    rows = ws.iter_rows(values_only=True)
    try:
        header = normalize_header(next(rows))
    except StopIteration:
        return []
    parsed: list[dict[str, Any]] = []
    for row in rows:
        if row is None:
            continue
        record = {}
        nonempty = False
        for key, idx in header.items():
            value = row[idx] if idx < len(row) else None
            record[key] = value
            if value is not None:
                nonempty = True
        if nonempty:
            parsed.append(record)
    return parsed


def safe_float(value: Any) -> float | None:
    try:
        if value is None:
            return None
        value = float(value)
        return None if math.isnan(value) else value
    except (TypeError, ValueError):
        return None


def resample_curve(voltage: list[float], capacity: list[float], max_points: int = 512) -> tuple[list[float], list[float]]:
    if len(voltage) <= max_points:
        return voltage, capacity

    cap_arr = np.asarray(capacity, dtype=float)
    volt_arr = np.asarray(voltage, dtype=float)
    order = np.argsort(cap_arr)
    cap_arr = cap_arr[order]
    volt_arr = volt_arr[order]
    uniq_cap, uniq_idx = np.unique(cap_arr, return_index=True)
    uniq_volt = volt_arr[uniq_idx]
    if len(uniq_cap) <= max_points:
        return uniq_volt.tolist(), uniq_cap.tolist()

    cap_grid = np.linspace(float(uniq_cap[0]), float(uniq_cap[-1]), max_points)
    volt_grid = np.interp(cap_grid, uniq_cap, uniq_volt)
    return volt_grid.tolist(), cap_grid.tolist()


def choose_key(header: dict[str, int], *candidates: str) -> int | None:
    for key in candidates:
        if key in header:
            return header[key]
    return None


def workbook_to_records(path: str) -> dict[str, Any]:
    path_obj = Path(path)
    batch = path_obj.parent.name
    cell_id = cell_id_from_name(path_obj.name)
    wb = load_workbook(path, read_only=True, data_only=True)

    if '循环' in wb.sheetnames:
        summary_name = '循环'
        cycle_key = '循环序号'
        charge_cap_key = '充电容量/mAh'
        discharge_cap_key = '放电容量/mAh'
        record_prefix = '记录'
        record_cols = ('循环序号', '电压/V', '电流/mA', '容量/mAh')
    elif 'Cycle' in wb.sheetnames:
        summary_name = 'Cycle'
        cycle_key = 'Cycle'
        charge_cap_key = 'CapC/mAh'
        discharge_cap_key = 'CapD/mAh'
        record_prefix = 'Record'
        record_cols = ('Cycle', 'Voltage/V', 'Current/mA', 'Capacity/mAh')
    else:
        return {'cell_id': cell_id, 'batch': batch, 'records': [], 'error': f'unsupported batch {batch}'}

    if summary_name not in wb.sheetnames:
        return {'cell_id': cell_id, 'batch': batch, 'records': [], 'error': f'missing summary sheet {summary_name}'}

    summary = read_sheet_rows(wb[summary_name])
    if not summary:
        return {'cell_id': cell_id, 'batch': batch, 'records': [], 'error': 'empty cycle sheet'}

    record_sheet_names = [sheet for sheet in wb.sheetnames if sheet.startswith(record_prefix)]
    if not record_sheet_names:
        return {'cell_id': cell_id, 'batch': batch, 'records': [], 'error': 'no record sheets'}

    discharge_points: dict[int, dict[str, list[float]]] = {}
    neg_current_seen = False
    sheet_headers: dict[str, list[str]] = {}

    for sheet_name in record_sheet_names:
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            continue
        header = normalize_header(header_row)
        sheet_headers[sheet_name] = list(header.keys())
        idx_cycle = choose_key(header, record_cols[0])
        idx_voltage = choose_key(header, record_cols[1])
        idx_current = choose_key(header, record_cols[2])
        idx_capacity = choose_key(header, record_cols[3])
        if None in (idx_cycle, idx_voltage, idx_current, idx_capacity):
            continue

        for row in rows:
            try:
                cycle = int(row[idx_cycle])
                voltage = float(row[idx_voltage])
                current = float(row[idx_current])
                capacity = float(row[idx_capacity])
            except (TypeError, ValueError):
                continue
            if math.isnan(current) or math.isnan(voltage) or math.isnan(capacity):
                continue
            if current < -1e-8:
                neg_current_seen = True
                store = discharge_points.setdefault(cycle, {'voltage': [], 'capacity': []})
                store['voltage'].append(voltage)
                store['capacity'].append(abs(capacity))

    if not neg_current_seen:
        return {
            'cell_id': cell_id,
            'batch': batch,
            'records': [],
            'error': 'no negative-current discharge rows',
            'sheet_headers': sheet_headers,
        }

    capacities: list[float | None] = []
    records: list[dict[str, Any]] = []
    for row in summary:
        try:
            cycle = int(row[cycle_key])
        except (TypeError, ValueError, KeyError):
            continue
        curve = discharge_points.get(cycle)
        if not curve or len(curve['voltage']) < 5:
            continue
        voltage_curve, capacity_curve = resample_curve(curve['voltage'], curve['capacity'])
        charge_cap = safe_float(row.get(charge_cap_key))
        discharge_cap = safe_float(row.get(discharge_cap_key))
        capacities.append(discharge_cap)
        records.append(
            {
                'cell_id': cell_id,
                'batch': batch,
                'cycle': cycle,
                'cycle_number': cycle,
                'charge_capacity': charge_cap,
                'discharge_capacity': discharge_cap,
                'voltage_discharge': voltage_curve,
                'capacity_discharge': capacity_curve,
            }
        )

    activation_end = None
    activation_flag = False
    valid_capacities = [cap for cap in capacities if cap is not None]
    if len(valid_capacities) >= 20 and records:
        series = pd.Series(valid_capacities, dtype='float64')
        rolling = series.rolling(window=5, min_periods=1).mean()
        peak_idx = int(rolling.idxmax())
        if peak_idx >= 4 and valid_capacities[0] > 0 and rolling.iloc[peak_idx] > valid_capacities[0] * 1.05:
            activation_flag = True
            activation_end = int(records[min(peak_idx, len(records) - 1)]['cycle_number'])

    for rec in records:
        if activation_flag and activation_end is not None and rec['cycle_number'] <= activation_end:
            rec['cycle_type'] = 'activation'
        else:
            rec['cycle_type'] = 'degradation'

    return {
        'cell_id': cell_id,
        'batch': batch,
        'records': records,
        'n_cycles': len(records),
        'path': str(path_obj),
        'sheet_headers': sheet_headers,
    }


def parse_local_files(paths: list[Path], workers: int = 4) -> tuple[pd.DataFrame, list[dict[str, Any]], list[str], dict[str, dict[str, list[str]]]]:
    rows: list[dict[str, Any]] = []
    cell_meta: list[dict[str, Any]] = []
    issues: list[str] = []
    header_examples: dict[str, dict[str, list[str]]] = {}

    with ProcessPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(workbook_to_records, str(path)): path for path in paths}
        for future in as_completed(futures):
            path = futures[future]
            try:
                result = future.result()
            except Exception as exc:
                issues.append(f'Failed to parse {path.name}: {exc}')
                continue
            batch = result.get('batch', path.parent.name)
            if result.get('sheet_headers') and not header_examples.get(batch):
                header_examples[batch] = result['sheet_headers']
            else:
                header_examples.setdefault(batch, result.get('sheet_headers', {}))
            if result.get('error'):
                issues.append(f"{path.name}: {result['error']}")
                continue
            rows.extend(result['records'])
            cell_meta.append(
                {
                    'cell_id': result['cell_id'],
                    'batch': batch,
                    'path': result['path'],
                    'n_cycles': result['n_cycles'],
                }
            )
    return pd.DataFrame(rows), cell_meta, issues, header_examples


def normalize_sequence(value: Any) -> list[float]:
    if value is None:
        return []
    if isinstance(value, list):
        return [float(x) for x in value]
    if hasattr(value, 'tolist'):
        value = value.tolist()
        if isinstance(value, list):
            return [float(x) for x in value]
    if isinstance(value, tuple):
        return [float(x) for x in value]
    return [float(x) for x in value]


def unique_cell_count(df: pd.DataFrame) -> int:
    if df.empty:
        return 0
    return int(df[['batch', 'cell_id']].drop_duplicates().shape[0])


def cycle_range(df: pd.DataFrame) -> list[int | None]:
    if df.empty:
        return [None, None]
    counts = df.groupby('cell_id')['cycle_number'].nunique()
    return [int(counts.min()), int(counts.max())]


def fetch_text(url: str) -> str:
    resp = _session().get(url, timeout=60)
    resp.raise_for_status()
    return resp.text


def scan_local_batch_temperature_strings(batch_dir: Path) -> list[str]:
    hits: list[str] = []
    temp_regex = re.compile(r'(温度|temperature|temp|℃|°C)', re.I)
    for path in sorted(batch_dir.glob('*.xlsx'))[:5]:
        try:
            with zipfile.ZipFile(path) as zf:
                for name in zf.namelist():
                    if not name.endswith('.xml'):
                        continue
                    text = zf.read(name).decode('utf-8', 'ignore')
                    if temp_regex.search(text):
                        hits.append(f'{path.name}:{name}')
                        break
        except Exception:
            continue
    return hits


def find_temperature_mapping() -> dict[str, Any]:
    stats_text = fetch_text(TEMP_STATS_URL)
    preprocess_text = fetch_text(TEMP_PREPROCESS_URL)
    aging_text = fetch_text(TEMP_AGING_URL)

    dataset_level_temps_known = all(token in stats_text for token in ('25', '30', '40')) and 'ZN-coin' in stats_text
    explicit_mapping_found = False
    mapping = {'Batch-1': 'unknown', 'Batch-2': 'unknown', 'Batch-3': 'unknown'}

    explicit_patterns = [
        re.compile(r'(Batch-[123]).{0,80}?(25|30|40)\s*(?:degrees Celsius|°C|℃)', re.I | re.S),
        re.compile(r'(25|30|40)\s*(?:degrees Celsius|°C|℃).{0,80}?(Batch-[123])', re.I | re.S),
    ]
    for text in (stats_text, preprocess_text, aging_text):
        for pattern in explicit_patterns:
            for match in pattern.finditer(text):
                groups = match.groups()
                batch = next((g for g in groups if g and g.startswith('Batch-')), None)
                temp = next((g for g in groups if g and g in {'25', '30', '40'}), None)
                if batch and temp:
                    mapping[batch] = f'{temp}°C'
                    explicit_mapping_found = True

    local_hits = {
        batch: scan_local_batch_temperature_strings(DATA_DIR / batch)
        for batch in ('Batch-1', 'Batch-2', 'Batch-3')
        if (DATA_DIR / batch).exists()
    }

    source = TEMP_STATS_URL if explicit_mapping_found else 'not found'
    return {
        'mapping': mapping,
        'temperature_mapping_found': explicit_mapping_found,
        'temperature_mapping_source': source,
        'dataset_level_temperatures_known': dataset_level_temps_known,
        'dataset_level_temperature_source': TEMP_STATS_URL if dataset_level_temps_known else 'not found',
        'local_temperature_string_hits': local_hits,
    }


def main() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    download_manifest: dict[str, list[dict[str, Any]]] = {}
    for batch in BATCHES_TO_DOWNLOAD:
        download_manifest[batch] = list_hf_files(batch)

    downloaded_paths: list[Path] = []
    with ThreadPoolExecutor(max_workers=4) as pool:
        futures = {
            pool.submit(download_one, item): item
            for items in download_manifest.values()
            for item in items
        }
        for future in as_completed(futures):
            downloaded_paths.append(future.result())

    downloaded_paths.sort()

    all_batch_paths = [
        path
        for batch in ('Batch-1', 'Batch-2', 'Batch-3')
        for path in sorted((DATA_DIR / batch).glob('*.xlsx'))
    ]
    combined, _, issues, header_examples = parse_local_files(all_batch_paths, workers=4)
    if not combined.empty:
        combined['voltage_discharge'] = combined['voltage_discharge'].apply(normalize_sequence)
        combined['capacity_discharge'] = combined['capacity_discharge'].apply(normalize_sequence)
    combined = combined.sort_values(['batch', 'cell_id', 'cycle_number']).reset_index(drop=True)

    PARQUET_PATH.parent.mkdir(parents=True, exist_ok=True)
    combined.to_parquet(PARQUET_PATH, index=False)

    temp_info = find_temperature_mapping()

    batch_frames = {batch: combined[combined['batch'] == batch].copy() for batch in ('Batch-1', 'Batch-2', 'Batch-3')}
    summary = {
        'batch_1': {
            'n_cells': unique_cell_count(batch_frames['Batch-1']),
            'temperature': temp_info['mapping'].get('Batch-1', 'unknown'),
            'cycle_range': cycle_range(batch_frames['Batch-1']),
        },
        'batch_2': {
            'n_cells': unique_cell_count(batch_frames['Batch-2']),
            'temperature': temp_info['mapping'].get('Batch-2', 'unknown'),
            'cycle_range': cycle_range(batch_frames['Batch-2']),
        },
        'batch_3': {
            'n_cells': unique_cell_count(batch_frames['Batch-3']),
            'temperature': temp_info['mapping'].get('Batch-3', 'unknown'),
            'cycle_range': cycle_range(batch_frames['Batch-3']),
        },
        'total_cells': unique_cell_count(combined),
        'temperature_mapping_found': temp_info['temperature_mapping_found'],
        'temperature_mapping_source': temp_info['temperature_mapping_source'],
    }

    with open(SUMMARY_PATH, 'w') as fh:
        json.dump(summary, fh, indent=2, ensure_ascii=False)

    print(json.dumps({
        'downloaded_files': len(downloaded_paths),
        'parsed_total_cells': unique_cell_count(combined),
        'batch_1_cells': summary['batch_1']['n_cells'],
        'batch_2_cells': summary['batch_2']['n_cells'],
        'batch_3_cells': summary['batch_3']['n_cells'],
        'total_cells': summary['total_cells'],
        'parquet': str(PARQUET_PATH),
        'summary': str(SUMMARY_PATH),
    }, indent=2, ensure_ascii=False))


if __name__ == '__main__':
    main()
