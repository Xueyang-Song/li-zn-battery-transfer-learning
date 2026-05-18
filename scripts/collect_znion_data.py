#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import math
import os
import re
import sys
import time
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any
from urllib.parse import quote

import numpy as np
import pandas as pd
import requests
from openpyxl import load_workbook

BASE_DIR = Path("/Users/melodysong/code/phd/battery_ml")
BATTERYLIFE_DIR = BASE_DIR / "data/zn_ion/batterylife"
ZENODO_DIR = BASE_DIR / "data/zn_ion/zenodo"
PARQUET_PATH = BASE_DIR / "processed/zn_ion_raw.parquet"
SUMMARY_PATH = BASE_DIR / "processed/znion_summary.json"
SEARCH_PATH = BASE_DIR / "processed/zn_ion_search_results.json"

sys.path.insert(0, str(BASE_DIR / "code"))
from pelt_activation import detect_activation_end  # noqa: E402


HF_TREE_URL = (
    "https://huggingface.co/api/datasets/Battery-Life/BatteryLife_Raw/"
    "tree/main?recursive=true&expand=false"
)
HF_FILE_BASE = "https://huggingface.co/datasets/Battery-Life/BatteryLife_Raw/resolve/main/"

BATTERYLIFE_DATASET = {
    "name": "BatteryLife_Raw ZNion",
    "source": "Hugging Face mirror of BatteryLife raw dataset",
    "url": "https://huggingface.co/datasets/Battery-Life/BatteryLife_Raw/tree/main/ZNion",
    "mirror_urls": [
        "https://zenodo.org/records/17960956",
        "https://github.com/Ruifeng-Tan/BatteryLife",
        "https://github.com/microsoft/BatteryML",
    ],
    "n_cells_reported": 95,
    "cathode_chemistry": "MnO2",
    "battery_type": "aqueous Zn-ion coin cells",
    "file_format": "XLSX",
    "cycles_per_cell_reported": "typically 500+ in selected files",
    "has_voltage_curves": True,
    "quality": 5,
    "notes": (
        "Paper reports first public Zn-ion life dataset; MnO2 cathode, Zn metal anode, "
        "three temperatures, one protocol, raw cycle summary and pointwise voltage/current/capacity."
    ),
}


NONUSABLE_CANDIDATES = [
    {
        "name": "ACS figshare Zn-V2O5 supplementary",
        "url": "https://acs.figshare.com/articles/journal_contribution/"
               "Rechargeable_Aqueous_Zn_V_sub_2_sub_O_sub_5_sub_Battery_with_High_"
               "Energy_Density_and_Long_Cycle_Life/6296897",
        "n_cells": None,
        "cathode_chemistry": "V2O5",
        "has_voltage_curves": False,
        "file_format": "PDF supplementary information",
        "cycles_per_cell": "single-paper figures only",
        "quality": 1,
        "notes": "Only SI PDF exposed via figshare API; no machine-readable full cycling curves.",
    },
    {
        "name": "ACS figshare γ-MnO2 supplementary",
        "url": "https://acs.figshare.com/articles/journal_contribution/"
               "Electrochemically_Induced_Structural_Transformation_in_a_MnO_sub_2_sub_"
               "Cathode_of_a_High_Capacity_Zinc_Ion_Battery_System/2164030",
        "n_cells": None,
        "cathode_chemistry": "MnO2",
        "has_voltage_curves": False,
        "file_format": "PDF supplementary information",
        "cycles_per_cell": "single-paper figures only",
        "quality": 1,
        "notes": "Only SI PDF exposed via figshare API; not suitable for ML-ready curve extraction.",
    },
]


def _session() -> requests.Session:
    sess = requests.Session()
    sess.headers.update({"User-Agent": "Mozilla/5.0"})
    return sess


def fetch_hf_tree() -> list[dict[str, Any]]:
    resp = _session().get(HF_TREE_URL, timeout=60)
    resp.raise_for_status()
    return resp.json()


def select_files(tree: list[dict[str, Any]], max_cells: int) -> list[dict[str, Any]]:
    files = [
        item for item in tree
        if item["type"] == "file"
        and item["path"].startswith("ZNion/")
        and item["path"].endswith(".xlsx")
    ]
    preferred: list[dict[str, Any]] = []
    for batch in ("Batch-2", "Batch-3", "Batch-1"):
        batch_files = [
            f for f in files
            if f["path"].split("/")[1] == batch and f.get("size", 0) > 8_000_000
        ]
        batch_files.sort(key=lambda x: (cell_sort_key(x["path"]), -x.get("size", 0)))
        preferred.extend(batch_files)
        if len(preferred) >= max_cells:
            break

    deduped: list[dict[str, Any]] = []
    seen: set[str] = set()
    for item in preferred:
        cell_id = extract_cell_id(Path(item["path"]).name)
        if cell_id in seen:
            continue
        seen.add(cell_id)
        deduped.append(item)
        if len(deduped) >= max_cells:
            break
    return deduped


def cell_sort_key(path: str) -> tuple[int, str]:
    cell = extract_cell_id(Path(path).name)
    nums = [int(x) for x in re.findall(r"\d+", cell)]
    return (nums[0] if nums else 999999, cell)


def extract_cell_id(name: str) -> str:
    dash_match = re.search(r"(\d+-\d+)", name)
    if dash_match:
        return dash_match.group(1)
    match = re.match(r"([^_]+)_", name)
    return match.group(1) if match else name.replace(".xlsx", "")


def download_one(item: dict[str, Any]) -> Path:
    rel_path = item["path"]
    out_path = BATTERYLIFE_DIR / rel_path
    out_path.parent.mkdir(parents=True, exist_ok=True)
    if out_path.exists() and out_path.stat().st_size == int(item.get("size", 0)):
        return out_path

    url = HF_FILE_BASE + quote(rel_path)
    sess = _session()
    for attempt in range(3):
        try:
            with sess.get(url, stream=True, timeout=120) as resp:
                resp.raise_for_status()
                with open(out_path, "wb") as fh:
                    for chunk in resp.iter_content(chunk_size=1024 * 1024):
                        if chunk:
                            fh.write(chunk)
            return out_path
        except Exception:
            if attempt == 2:
                raise
            time.sleep(2 + attempt)
    return out_path


def workbook_to_records(path: str) -> dict[str, Any]:
    path_obj = Path(path)
    cell_id = extract_cell_id(path_obj.name)
    wb = load_workbook(path, read_only=True, data_only=True)
    summary = read_sheet_rows(wb["循环"])
    if not summary:
        return {"cell_id": cell_id, "records": [], "error": "empty cycle sheet"}

    record_sheet_names = [s for s in wb.sheetnames if s.startswith("记录")]
    discharge_points: dict[int, dict[str, list[float]]] = {}
    neg_current_seen = False

    for sheet_name in record_sheet_names:
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        header = normalize_header(next(rows))
        idx_cycle = header.get("循环序号")
        idx_voltage = header.get("电压/V")
        idx_current = header.get("电流/mA")
        idx_capacity = header.get("容量/mAh")
        if None in (idx_cycle, idx_voltage, idx_current, idx_capacity):
            continue

        for row in rows:
            try:
                cycle = int(row[idx_cycle])
                voltage = float(row[idx_voltage])
                current = float(row[idx_current])
                capacity = float(row[idx_capacity])
            except (TypeError, ValueError):
                continue
            if math.isnan(current) or math.isnan(voltage) or math.isnan(capacity):
                continue
            if current < -1e-8:
                neg_current_seen = True
                store = discharge_points.setdefault(
                    cycle, {"voltage": [], "capacity": []}
                )
                store["voltage"].append(voltage)
                store["capacity"].append(abs(capacity))

    if not neg_current_seen:
        return {"cell_id": cell_id, "records": [], "error": "no negative-current discharge rows"}

    capacities = []
    records: list[dict[str, Any]] = []
    for row in summary:
        try:
            cycle = int(row["循环序号"])
        except (TypeError, ValueError, KeyError):
            continue
        curve = discharge_points.get(cycle)
        if not curve or len(curve["voltage"]) < 5:
            continue
        voltage_curve, capacity_curve = resample_curve(
            curve["voltage"], curve["capacity"], max_points=512
        )
        charge_cap = safe_float(row.get("充电容量/mAh"))
        discharge_cap = safe_float(row.get("放电容量/mAh"))
        capacities.append(discharge_cap)
        records.append(
            {
                "cell_id": cell_id,
                "cycle": cycle,
                "cycle_number": cycle,
                "charge_capacity": charge_cap,
                "discharge_capacity": discharge_cap,
                "voltage_discharge": voltage_curve,
                "capacity_discharge": capacity_curve,
            }
        )

    activation_end = None
    activation_flag = False
    if len(capacities) >= 20:
        try:
            activation_end_idx, info = detect_activation_end(pd.Series(capacities).to_numpy())
            activation_end = records[activation_end_idx]["cycle_number"] if records else None
            first = capacities[0]
            peak_window = max(capacities[: min(20, len(capacities))])
            activation_flag = (
                activation_end_idx >= 4 and first > 0 and peak_window > first * 1.05
            )
        except Exception:
            activation_end = None

    for rec in records:
        if activation_flag and activation_end is not None and rec["cycle_number"] <= activation_end:
            rec["cycle_type"] = "activation"
        else:
            rec["cycle_type"] = "degradation"

    return {
        "cell_id": cell_id,
        "records": records,
        "n_cycles": len(records),
        "activation_flag": activation_flag,
        "activation_end_cycle": activation_end,
        "path": str(path_obj),
    }


def normalize_header(row: Any) -> dict[str, int]:
    header: dict[str, int] = {}
    for idx, value in enumerate(row):
        if value is None:
            continue
        header[str(value).strip()] = idx
    return header


def read_sheet_rows(ws) -> list[dict[str, Any]]:
    rows = ws.iter_rows(values_only=True)
    header = normalize_header(next(rows))
    parsed: list[dict[str, Any]] = []
    for row in rows:
        if row is None:
            continue
        record = {}
        nonempty = False
        for key, idx in header.items():
            value = row[idx] if idx < len(row) else None
            record[key] = value
            if value is not None:
                nonempty = True
        if nonempty:
            parsed.append(record)
    return parsed


def safe_float(value: Any) -> float | None:
    try:
        if value is None:
            return None
        value = float(value)
        return None if math.isnan(value) else value
    except (TypeError, ValueError):
        return None


def resample_curve(
    voltage: list[float], capacity: list[float], max_points: int = 512
) -> tuple[list[float], list[float]]:
    if len(voltage) <= max_points:
        return voltage, capacity

    cap_arr = np.asarray(capacity, dtype=float)
    volt_arr = np.asarray(voltage, dtype=float)
    order = np.argsort(cap_arr)
    cap_arr = cap_arr[order]
    volt_arr = volt_arr[order]
    uniq_cap, uniq_idx = np.unique(cap_arr, return_index=True)
    uniq_volt = volt_arr[uniq_idx]
    if len(uniq_cap) <= max_points:
        return uniq_volt.tolist(), uniq_cap.tolist()

    cap_grid = np.linspace(float(uniq_cap[0]), float(uniq_cap[-1]), max_points)
    volt_grid = np.interp(cap_grid, uniq_cap, uniq_volt)
    return volt_grid.tolist(), cap_grid.tolist()


def parse_local_files(paths: list[Path], workers: int) -> tuple[pd.DataFrame, list[dict[str, Any]], list[str]]:
    rows: list[dict[str, Any]] = []
    cell_meta: list[dict[str, Any]] = []
    issues: list[str] = []

    with ProcessPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(workbook_to_records, str(path)): path for path in paths}
        for future in as_completed(futures):
            path = futures[future]
            try:
                result = future.result()
            except Exception as exc:
                issues.append(f"Failed to parse {path.name}: {exc}")
                continue
            if result.get("error"):
                issues.append(f"{path.name}: {result['error']}")
                continue
            rows.extend(result["records"])
            cell_meta.append(
                {
                    "cell_id": result["cell_id"],
                    "path": result["path"],
                    "n_cycles": result["n_cycles"],
                    "activation_flag": result["activation_flag"],
                    "activation_end_cycle": result["activation_end_cycle"],
                }
            )
    df = pd.DataFrame(rows)
    return df, cell_meta, issues


def build_summary(
    df_all: pd.DataFrame,
    df_qc: pd.DataFrame,
    cell_meta: list[dict[str, Any]],
    issues: list[str],
    selected_files: list[dict[str, Any]],
) -> dict[str, Any]:
    known_issues = [
        "Zenodo search and record access from this environment returned HTTP 403 (unusual traffic block), so the Hugging Face mirror was used for BatteryLife raw downloads.",
        "Harvard Dataverse search did not surface any relevant Zn-ion full-curve life dataset.",
        "GitHub and Hugging Face searches did not surface any second public Zn-ion life dataset with machine-readable full voltage curves.",
        "Figshare V2O5/MnO2 zinc-battery hits exposed only supplementary PDFs, not downloadable full cycling tables.",
        "Publicly found ML-ready Zn-ion data remain MnO2-only; the ≥2 cathode-chemistry target is still unmet.",
    ]
    merged_issues = known_issues + issues
    n_cells_total = int(df_all["cell_id"].nunique()) if not df_all.empty else 0
    n_cells_qc = int(df_qc["cell_id"].nunique()) if not df_qc.empty else 0
    activation_cells = sum(1 for meta in cell_meta if meta["activation_flag"])

    datasets_found = [
        {
            **BATTERYLIFE_DATASET,
            "n_cells_downloaded": len(selected_files),
            "n_cells_parsed": n_cells_total,
            "cycles_per_cell_observed": summarize_cycle_counts(cell_meta),
        },
        *NONUSABLE_CANDIDATES,
    ]

    return {
        "datasets_found": datasets_found,
        "datasets_downloaded": [BATTERYLIFE_DATASET["name"]],
        "downloaded_files": [item["path"] for item in selected_files],
        "n_cells_total": n_cells_total,
        "n_cells_with_voltage": n_cells_total,
        "n_cells_after_qc": n_cells_qc,
        "cathode_types": ["MnO2"],
        "cells_with_activation_period": activation_cells,
        "issues": merged_issues,
        "recommendation": (
            "Use BatteryLife_Raw ZNion (selected long-cycling XLSX files with full "
            "voltage curves). It is strong enough for MnO2-only transfer experiments, "
            "but the public search did not uncover a second ML-ready V2O5/MnO2 dataset, "
            "so the cathode-diversity target remains unmet."
        ),
    }


def summarize_cycle_counts(cell_meta: list[dict[str, Any]]) -> dict[str, Any]:
    if not cell_meta:
        return {"min": None, "median": None, "max": None}
    counts = sorted(meta["n_cycles"] for meta in cell_meta)
    return {
        "min": counts[0],
        "median": counts[len(counts) // 2],
        "max": counts[-1],
    }


def save_outputs(df_qc: pd.DataFrame, summary: dict[str, Any]) -> None:
    PARQUET_PATH.parent.mkdir(parents=True, exist_ok=True)
    SUMMARY_PATH.parent.mkdir(parents=True, exist_ok=True)
    df_qc.to_parquet(PARQUET_PATH, index=False)
    with open(SUMMARY_PATH, "w") as fh:
        json.dump(summary, fh, indent=2, ensure_ascii=False)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--max-cells", type=int, default=40)
    parser.add_argument("--download-workers", type=int, default=4)
    parser.add_argument("--parse-workers", type=int, default=4)
    args = parser.parse_args()

    BATTERYLIFE_DIR.mkdir(parents=True, exist_ok=True)
    ZENODO_DIR.mkdir(parents=True, exist_ok=True)

    tree = fetch_hf_tree()
    selected = select_files(tree, max_cells=args.max_cells)

    local_paths: list[Path] = []
    with ThreadPoolExecutor(max_workers=args.download_workers) as pool:
        futures = {pool.submit(download_one, item): item for item in selected}
        for future in as_completed(futures):
            item = futures[future]
            path = future.result()
            local_paths.append(path)
            print(f"downloaded {item['path']} -> {path}")

    local_paths.sort(key=lambda p: cell_sort_key(str(p)))
    df_all, cell_meta, issues = parse_local_files(local_paths, workers=args.parse_workers)

    if df_all.empty:
        raise RuntimeError("No Zn-ion records were parsed.")

    valid_cells = {
        meta["cell_id"] for meta in cell_meta
        if meta["n_cycles"] >= 50
    }
    df_qc = df_all[df_all["cell_id"].isin(valid_cells)].copy()

    summary = build_summary(df_all, df_qc, cell_meta, issues, selected)
    save_outputs(df_qc, summary)

    print(json.dumps(
        {
            "downloaded_cells": len(local_paths),
            "parsed_cells": df_all["cell_id"].nunique(),
            "qc_cells": df_qc["cell_id"].nunique(),
            "rows_saved": len(df_qc),
            "parquet": str(PARQUET_PATH),
            "summary": str(SUMMARY_PATH),
        },
        indent=2,
        ensure_ascii=False,
    ))


if __name__ == "__main__":
    main()
