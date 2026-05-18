"""BatteryLife Zn-ion data loader.

Loads pre-downloaded BatteryLife XLSX files from the local ``data/zn_ion/batterylife``
directory.  Each XLSX file contains cycle-summary and pointwise voltage/current/capacity
data for one cell.

Refactored from ``scripts/collect_znion_data.py`` into a purely functional module.
"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any, Optional

import numpy as np
import pandas as pd

from config.settings import PipelineSettings

logger = logging.getLogger(__name__)

# Chinese column name mappings (BatteryLife uses Chinese headers)
_COL_MAP = {
    "循环序号": "cycle",
    "放电容量/mAh": "discharge_capacity",
    "充电容量/mAh": "charge_capacity",
    "电压/V": "voltage",
    "电流/mA": "current",
    "容量/mAh": "capacity",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _extract_cell_id(filename: str) -> str:
    """Derive a clean cell ID from an XLSX filename.

    Args:
        filename: XLSX file name (may contain batch/cell numbering patterns).

    Returns:
        Short cell ID string, e.g. ``"12-3"`` or ``"cell_001"``.
    """
    dash_match = re.search(r"(\d+-\d+)", filename)
    if dash_match:
        return dash_match.group(1)
    match = re.match(r"([^_]+)_", filename)
    if match:
        return match.group(1)
    return filename.replace(".xlsx", "")


def _normalize_header(row: tuple) -> dict[str, int]:
    """Map Chinese header names to column indices.

    Args:
        row: First row of the worksheet (values_only=True).

    Returns:
        Dict mapping known Chinese column names to their 0-based indices.
    """
    return {
        cell: idx
        for idx, cell in enumerate(row)
        if cell in _COL_MAP
    }


def _read_cycle_summary(
    xlsx_path: Path,
) -> Optional[pd.DataFrame]:
    """Parse the cycle-summary sheet (``循环``) from a BatteryLife XLSX.

    Args:
        xlsx_path: Path to the XLSX file.

    Returns:
        DataFrame with columns ``cycle`` and ``discharge_capacity``, or
        ``None`` if the sheet cannot be read.
    """
    try:
        from openpyxl import load_workbook  # lazy import
        wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        logger.warning("Cannot open XLSX", extra={"path": str(xlsx_path), "error": str(exc)})
        return None

    if "循环" not in wb.sheetnames:
        logger.warning("Cycle sheet missing", extra={"path": str(xlsx_path)})
        wb.close()
        return None

    ws = wb["循环"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return None

    header = _normalize_header(rows[0])
    idx_cycle = header.get("循环序号")
    idx_cap = header.get("放电容量/mAh")

    if idx_cycle is None or idx_cap is None:
        # Try alternate column names
        for i, col in enumerate(rows[0]):
            if col and "序号" in str(col):
                idx_cycle = i
            if col and "放电" in str(col) and "容量" in str(col):
                idx_cap = i

    if idx_cycle is None or idx_cap is None:
        logger.warning(
            "Required columns not found in cycle sheet",
            extra={"path": str(xlsx_path), "headers": list(rows[0])},
        )
        return None

    cycles, capacities = [], []
    for row in rows[1:]:
        try:
            cyc = int(row[idx_cycle])
            cap = float(row[idx_cap])
            if cap > 0:
                cycles.append(cyc)
                capacities.append(cap)
        except (TypeError, ValueError):
            continue

    if not cycles:
        return None

    return pd.DataFrame({"cycle": cycles, "discharge_capacity": capacities})


# ---------------------------------------------------------------------------
# Per-cell feature extraction
# ---------------------------------------------------------------------------

def _extract_cell_record(
    xlsx_path: Path,
    batch: str,
    settings: PipelineSettings,
) -> Optional[dict[str, Any]]:
    """Extract a cell record from one BatteryLife XLSX file.

    Args:
        xlsx_path: Path to the XLSX file.
        batch: Batch identifier (e.g. ``"Batch-1"``).
        settings: Pipeline configuration.

    Returns:
        Dict with cell metadata and capacity curve, or ``None`` on failure.
    """
    raw_cell_id = _extract_cell_id(xlsx_path.name)
    cell_id = f"batterylife_{batch}_{raw_cell_id}"

    df = _read_cycle_summary(xlsx_path)
    if df is None or len(df) < settings.min_cycle_life:
        logger.debug(
            "Cell skipped (too short or no data)",
            extra={"cell_id": cell_id, "n_cycles": len(df) if df is not None else 0},
        )
        return None

    df = df.sort_values("cycle").reset_index(drop=True)
    capacity = df["discharge_capacity"].values.astype(float)

    # Cycle life: first cycle below 80% of max capacity
    q_max = float(np.max(capacity))
    if q_max <= 0:
        return None

    q_norm = capacity / q_max
    below = np.where(q_norm < 0.80)[0]
    cycle_life = int(below[0]) if len(below) > 0 else len(capacity)

    if cycle_life < settings.min_cycle_life:
        return None

    return {
        "cell_id": cell_id,
        "dataset": "batterylife",
        "chemistry": "ZnMnO2",
        "batch": batch,
        "cycle_life": float(cycle_life),
        "n_cycles_raw": int(len(capacity)),
        "raw_capacity_curve": capacity.tolist(),
    }


# ---------------------------------------------------------------------------
# Public loader
# ---------------------------------------------------------------------------

def load_batterylife(
    data_dir: Path,
    settings: PipelineSettings,
) -> pd.DataFrame:
    """Load all available BatteryLife Zn-ion cells from local XLSX files.

    Searches recursively for ``*.xlsx`` files under ``data_dir``.  Files are
    grouped by batch directory name.

    Args:
        data_dir: Root directory containing batch sub-directories (e.g.
            ``ZNion/Batch-1/``, ``ZNion/Batch-2/``, …).
        settings: Pipeline configuration.

    Returns:
        DataFrame with one row per accepted cell.  Returns an empty DataFrame
        with a warning if ``data_dir`` does not exist.
    """
    if not data_dir.exists():
        logger.warning(
            "BatteryLife data directory not found — returning empty DataFrame",
            extra={"path": str(data_dir)},
        )
        return pd.DataFrame()

    xlsx_files = sorted(data_dir.rglob("*.xlsx"))
    if not xlsx_files:
        logger.warning("No XLSX files found", extra={"dir": str(data_dir)})
        return pd.DataFrame()

    records: list[dict] = []
    n_fail = 0

    for xlsx_path in xlsx_files:
        # Derive batch from parent directory name
        batch = xlsx_path.parent.name
        rec = _extract_cell_record(xlsx_path, batch, settings)
        if rec is None:
            n_fail += 1
        else:
            records.append(rec)

    logger.info(
        "BatteryLife loading complete",
        extra={"n_loaded": len(records), "n_failed": n_fail},
    )
    return pd.DataFrame(records)
